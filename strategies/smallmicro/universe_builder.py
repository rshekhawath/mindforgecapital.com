"""
build_stock_universe.py
=======================
One-stop script to build the master stock universe for the
MultiFactor SmallMicro 500 portfolio.

What it does
------------
  1. Fetches Nifty Smallcap 250 constituents directly from NSE
  2. Fetches Nifty Microcap 250 constituents directly from NSE
  3. Strips any NSE dummy/placeholder rows (e.g. DUMMYALCAR)
  4. Deduplicates — if a symbol appears in both indexes, it is
     tagged "Smallcap 250 & Microcap 250" and kept once
  5. Writes a single Excel file with one sheet:

     Col A  →  NSE Symbol             (e.g. INFY)
     Col B  →  Yahoo Finance Ticker   (e.g. INFY.NS)
     Col C  →  Industry               (NSE official classification)
     Col D  →  Index                  (Smallcap 250 / Microcap 250)
     Col E  →  Company Name           (full name from NSE)

No existing Excel file is required to run this script.
The output file is what source_code.py and backtest.py import.

Requirements
------------
    pip install pandas openpyxl requests

Run
---
    python build_stock_universe.py
"""

import os
import sys
import requests
import pandas as pd
from io import StringIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG  —  only change these two lines if your folder path differs
# ─────────────────────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(BASE_DIR, "universe.xlsx")

# NSE official index constituent CSV endpoints
NSE_INDICES = {
    "Smallcap 250": "https://nsearchives.nseindia.com/content/indices/ind_niftysmallcap250list.csv",
    "Microcap 250": "https://nsearchives.nseindia.com/content/indices/ind_niftymicrocap250_list.csv",
}

# Headers required to pass NSE's bot-detection
NSE_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Referer":         "https://www.nseindia.com/",
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
}
# ─────────────────────────────────────────────────────────────────────────────


# ── FETCH ─────────────────────────────────────────────────────────────────────
def fetch_index(session, name, url):
    """
    Download one NSE index CSV and return a clean DataFrame.

    NSE CSV format:
        Company Name | Industry | Symbol | Series | ISIN Code
    """
    print(f"  Fetching {name} ...", end=" ", flush=True)
    response = session.get(url, headers=NSE_HEADERS, timeout=20)
    response.raise_for_status()

    df = pd.read_csv(StringIO(response.text))
    df.columns = df.columns.str.strip()

    # Validate that the expected columns are present
    for col in ("Symbol", "Company Name", "Industry"):
        if col not in df.columns:
            raise ValueError(
                f"NSE CSV for '{name}' is missing column '{col}'. "
                f"Columns found: {list(df.columns)}"
            )

    df["Symbol"]       = df["Symbol"].str.strip().str.upper()
    df["Company Name"] = df["Company Name"].str.strip()
    df["Industry"]     = df["Industry"].str.strip()

    print(f"✓  ({len(df)} stocks)")
    return df[["Symbol", "Company Name", "Industry"]]


# ── STYLE ─────────────────────────────────────────────────────────────────────
def style_sheet(ws):
    """Apply professional dark-header formatting to a worksheet."""
    HEADER_FILL = PatternFill("solid", fgColor="1F3864")   # dark navy
    ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")   # light grey

    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    BODY_FONT   = Font(name="Arial", size=10)

    THIN   = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT   = Alignment(horizontal="left",   vertical="center")

    # Column widths: A=Symbol, B=YF Ticker, C=Industry, D=Index, E=Company Name
    col_widths = {1: 18, 2: 22, 3: 36, 4: 26, 5: 42}
    for col_num, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        is_header = row_idx == 1
        for cell in row:
            cell.border = BORDER
            if is_header:
                cell.fill      = HEADER_FILL
                cell.font      = HEADER_FONT
                cell.alignment = CENTER
            else:
                if row_idx % 2 == 0:
                    cell.fill = ALT_FILL
                cell.font      = BODY_FONT
                # Centre A (Symbol), B (YF Ticker), D (Index); left-align C and E
                cell.alignment = CENTER if cell.column in (1, 2, 4) else LEFT

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    print("=" * 62)
    print("  NSE Universe Builder — Nifty Smallcap 250 + Microcap 250")
    print("=" * 62)

    # ── Step 1: Fetch both indexes from NSE ───────────────────────────
    print("\n[1] Fetching live constituent data from NSE...")
    session = requests.Session()

    # Warm-up GET to obtain the session cookie NSE requires
    try:
        session.get("https://www.nseindia.com", headers=NSE_HEADERS, timeout=10)
    except Exception as exc:
        print(f"  Warning: NSE warm-up request failed ({exc}). Continuing anyway.")

    index_dfs = {}
    for name, url in NSE_INDICES.items():
        try:
            index_dfs[name] = fetch_index(session, name, url)
        except Exception as exc:
            print(f"\n  ERROR fetching '{name}': {exc}")
            print("  Check your internet connection and try again.")
            sys.exit(1)

    # ── Step 2: Combine and clean ─────────────────────────────────────
    print("\n[2] Building master universe...")

    combined = pd.concat(index_dfs.values(), ignore_index=True)

    # Remove NSE dummy/placeholder symbols (e.g. DUMMYALCAR).
    # NSE temporarily inserts these when a stock exits an index between
    # the semi-annual rebalance dates.
    dummy_mask = combined["Symbol"].str.upper().str.startswith("DUMMY")
    if dummy_mask.any():
        removed  = combined.loc[dummy_mask, "Symbol"].tolist()
        combined = combined[~dummy_mask].reset_index(drop=True)
        print(f"  Removed {len(removed)} NSE dummy placeholder(s): {removed}")

    # Build per-index symbol sets (after dummy removal) for index tagging
    sc_syms = set(index_dfs["Smallcap 250"]["Symbol"].str.upper())
    mc_syms = set(index_dfs["Microcap 250"]["Symbol"].str.upper())
    # Remove any dummies that appeared in the original sets
    sc_syms = {s for s in sc_syms if not s.startswith("DUMMY")}
    mc_syms = {s for s in mc_syms if not s.startswith("DUMMY")}

    def index_tag(sym):
        in_sc = sym in sc_syms
        in_mc = sym in mc_syms
        if in_sc and in_mc:
            return "Smallcap 250 & Microcap 250"
        return "Smallcap 250" if in_sc else "Microcap 250"

    # Deduplicate — if a symbol appears in both CSVs keep one row
    # (Smallcap 250 is fetched first so its Industry label wins on conflict)
    combined = combined.drop_duplicates(subset="Symbol", keep="first").reset_index(drop=True)

    # Add derived columns
    combined["Yahoo Finance Ticker"] = combined["Symbol"] + ".NS"
    combined["Index"]                = combined["Symbol"].map(index_tag)

    # Final column order — matches exactly what source_code.py and backtest.py expect
    master = combined[[
        "Symbol",
        "Yahoo Finance Ticker",
        "Industry",
        "Index",
        "Company Name",
    ]].sort_values(["Index", "Symbol"]).reset_index(drop=True)

    both_count = (master["Index"] == "Smallcap 250 & Microcap 250").sum()
    print(f"  Total stocks    : {len(master)}")
    print(f"  Smallcap 250    : {(master['Index'] == 'Smallcap 250').sum()}")
    print(f"  Microcap 250    : {(master['Index'] == 'Microcap 250').sum()}")
    if both_count:
        print(f"  In both indexes : {both_count}")
    print(f"  Industries      : {master['Industry'].nunique()} — "
          f"{', '.join(sorted(master['Industry'].unique()))}")

    # ── Step 3: Write Excel ───────────────────────────────────────────
    print(f"\n[3] Writing Excel to:\n    {OUTPUT_PATH}")
    os.makedirs(BASE_DIR, exist_ok=True)

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        master.to_excel(writer, sheet_name="Master Universe", index=False)

    # Apply formatting
    wb = load_workbook(OUTPUT_PATH)
    style_sheet(wb["Master Universe"])
    wb.save(OUTPUT_PATH)

    # ── Step 4: Summary ───────────────────────────────────────────────
    print("\n" + "=" * 62)
    print("  Done.")
    print(f"\n  Output : {OUTPUT_PATH}")
    print( "  Sheet  : Master Universe")
    print( "  Columns: Symbol | Yahoo Finance Ticker | Industry | Index | Company Name")
    print("\n  Use these settings in portfolio_builder.py and backtest.py:")
    print(f'  UNIVERSE_PATH  = r"{OUTPUT_PATH}"')
    print( '  UNIVERSE_SHEET = "Master Universe"')
    print("=" * 62 + "\n")


if __name__ == "__main__":
    main()