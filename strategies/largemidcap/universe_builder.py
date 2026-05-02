"""
build_largemid_universe.py
==========================
One-stop script to build the master stock universe for the
MultiFactor LargeMidcap 250 portfolio.

What it does
------------
  1. Fetches Nifty LargeMidcap 250 constituents directly from NSE
  2. Strips any NSE dummy/placeholder rows (e.g. DUMMY*)
  3. Writes a single Excel file with one sheet "Master Universe":

     Col A  →  NSE Symbol             (e.g. INFY)
     Col B  →  Yahoo Finance Ticker   (e.g. INFY.NS)
     Col C  →  Industry               (NSE official classification)
     Col D  →  Company Name           (full name from NSE)

No existing Excel file is required to run this script.
The output is what source_code.py and backtest.py import.

Requirements
------------
    pip install pandas openpyxl requests

Run
---
    python build_largemid_universe.py
"""

import os
import sys
import requests
import pandas as pd
from io import StringIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(BASE_DIR, "universe.xlsx")

# NSE official Nifty LargeMidcap 250 constituent CSV
NSE_URL = "https://nsearchives.nseindia.com/content/indices/ind_niftylargemidcap250list.csv"

NSE_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Referer":         "https://www.nseindia.com/",
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
}
# ─────────────────────────────────────────────────────────────────────────────


def fetch_index(session):
    """
    Download Nifty LargeMidcap 250 CSV and return a clean DataFrame.

    NSE CSV format:
        Company Name | Industry | Symbol | Series | ISIN Code
    """
    print("  Fetching Nifty LargeMidcap 250 ...", end=" ", flush=True)
    response = session.get(NSE_URL, headers=NSE_HEADERS, timeout=20)
    response.raise_for_status()

    df = pd.read_csv(StringIO(response.text))
    df.columns = df.columns.str.strip()

    for col in ("Symbol", "Company Name", "Industry"):
        if col not in df.columns:
            raise ValueError(
                f"NSE CSV is missing column '{col}'. "
                f"Columns found: {list(df.columns)}"
            )

    df["Symbol"]       = df["Symbol"].str.strip().str.upper()
    df["Company Name"] = df["Company Name"].str.strip()
    df["Industry"]     = df["Industry"].str.strip()

    print(f"✓  ({len(df)} stocks)")
    return df[["Symbol", "Company Name", "Industry"]]


def style_sheet(ws):
    """Apply professional dark-header formatting to a worksheet."""
    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    BODY_FONT   = Font(name="Arial", size=10)
    THIN        = Side(style="thin", color="D9D9D9")
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER      = Alignment(horizontal="center", vertical="center")
    LEFT        = Alignment(horizontal="left",   vertical="center")

    # A=Symbol, B=YF Ticker, C=Industry, D=Company Name
    col_widths = {1: 18, 2: 22, 3: 38, 4: 46}
    for col_num, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        is_header = row_idx == 1
        for cell in row:
            cell.border = BORDER
            if is_header:
                cell.fill      = HEADER_FILL
                cell.font      = HEADER_FONT
                cell.alignment = CENTER
            else:
                if row_idx % 2 == 0:
                    cell.fill = ALT_FILL
                cell.font      = BODY_FONT
                # Centre A (Symbol) and B (YF Ticker); left-align C and D
                cell.alignment = CENTER if cell.column in (1, 2) else LEFT

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main():
    print("=" * 62)
    print("  NSE Universe Builder — Nifty LargeMidcap 250")
    print("=" * 62)

    # ── Step 1: Fetch from NSE ────────────────────────────────────────
    print("\n[1] Fetching live constituent data from NSE...")
    session = requests.Session()
    try:
        session.get("https://www.nseindia.com", headers=NSE_HEADERS, timeout=10)
    except Exception as exc:
        print(f"  Warning: NSE warm-up failed ({exc}). Continuing anyway.")

    try:
        df = fetch_index(session)
    except Exception as exc:
        print(f"\n  ERROR: {exc}")
        print("  Check your internet connection and try again.")
        sys.exit(1)

    # ── Step 2: Clean ─────────────────────────────────────────────────
    print("\n[2] Building master universe...")

    # Remove NSE dummy/placeholder rows
    dummy_mask = df["Symbol"].str.upper().str.startswith("DUMMY")
    if dummy_mask.any():
        removed = df.loc[dummy_mask, "Symbol"].tolist()
        df      = df[~dummy_mask].reset_index(drop=True)
        print(f"  Removed {len(removed)} NSE dummy placeholder(s): {removed}")

    # Add Yahoo Finance ticker
    df["Yahoo Finance Ticker"] = df["Symbol"] + ".NS"

    # Final column order
    master = df[[
        "Symbol",
        "Yahoo Finance Ticker",
        "Industry",
        "Company Name",
    ]].sort_values("Symbol").reset_index(drop=True)

    print(f"  Total stocks : {len(master)}")
    print(f"  Industries   : {master['Industry'].nunique()} — "
          f"{', '.join(sorted(master['Industry'].unique()))}")

    # ── Step 3: Write Excel ───────────────────────────────────────────
    print(f"\n[3] Writing Excel to:\n    {OUTPUT_PATH}")
    os.makedirs(BASE_DIR, exist_ok=True)

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        master.to_excel(writer, sheet_name="Master Universe", index=False)

    wb = load_workbook(OUTPUT_PATH)
    style_sheet(wb["Master Universe"])
    wb.save(OUTPUT_PATH)

    # ── Step 4: Summary ───────────────────────────────────────────────
    print("\n" + "=" * 62)
    print("  Done.")
    print(f"\n  Output : {OUTPUT_PATH}")
    print( "  Sheet  : Master Universe")
    print( "  Columns: Symbol | Yahoo Finance Ticker | Industry | Company Name")
    print("\n  Use these settings in portfolio_builder.py and backtest.py:")
    print(f'  UNIVERSE_PATH  = r"{OUTPUT_PATH}"')
    print( '  UNIVERSE_SHEET = "Master Universe"')
    print("=" * 62 + "\n")


if __name__ == "__main__":
    main()