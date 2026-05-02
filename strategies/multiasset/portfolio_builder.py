"""
MultiFactor MultiAsset — Source Script
========================================
Universe: 6 assets (Nifty 50, Nifty Next 50, Nifty Midcap 150,
          Gold, Bharat Bond, NASDAQ)
"""

import os
import sys
import yfinance as yf
import pandas as pd
import numpy as np
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =========================================================
# PATHS  (all relative — works wherever the repo lives)
# =========================================================

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
OUTPUTS_DIR = os.path.join(BASE_DIR, "outputs")
os.makedirs(OUTPUTS_DIR, exist_ok=True)

OUTPUT_FILE = os.path.join(OUTPUTS_DIR, "weekly_momentum_table.xlsx")
LOG_FILE    = os.path.join(OUTPUTS_DIR, "run.log")

ASSETS = {
    "Nifty 50 ETF":         "NIFTYBEES.NS",
    "Nifty Next 50 ETF":    "JUNIORBEES.NS",
    "Nifty Midcap 150 ETF": "MID150BEES.NS",
    "Gold ETF":             "GOLDCASE.NS",
    "Bharat Bond ETF":      "EBBETF0431.NS",
    "NASDAQ ETF":           "MON100.NS",
}

# Lookback windows (weeks)
WEEKS_3M  = 13
WEEKS_6M  = 26
WEEKS_12M = 52

# Trend filter: 10-month moving average (~40 weekly bars)
MA_10_MONTHS = 40

# ── Volatility penalty ────────────────────────────────────
VOL_PENALTY   = 0.5
VOL_ADJ_FLOOR = 0.2

# ── Trend penalty ─────────────────────────────────────────
TREND_PENALTY = 0.5

# =========================================================
# LOGGING
# =========================================================

logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logging.info("===== RUN START =====")

# =========================================================
# HELPERS
# =========================================================

def momentum(current, past):
    return (current / past) - 1


def annualised_vol(weekly_closes):
    """52-week annualised volatility from weekly returns."""
    rets = weekly_closes.pct_change().dropna()
    return rets.std() * np.sqrt(52)


# =========================================================
# FETCH & COMPUTE
# =========================================================

asset_data = {}   # name → dict of computed values

for asset_name, ticker in ASSETS.items():
    try:
        logging.info(f"Fetching {asset_name} ({ticker})")

        tk = yf.Ticker(ticker)
        df = tk.history(period="3y", interval="1d", auto_adjust=True)

        if df.empty:
            logging.error(f"{asset_name}: EMPTY DATA FROM YAHOO")
            continue

        df = df.resample("W-FRI").last().dropna()
        logging.info(f"{asset_name}: {len(df)} weekly rows")

        if len(df) < WEEKS_12M + 1:
            logging.error(f"{asset_name}: NOT ENOUGH DATA ({len(df)} rows)")
            continue

        close = df["Close"]
        current = close.iloc[-1]

        # Momentum
        p3  = close.iloc[-WEEKS_3M]
        p6  = close.iloc[-WEEKS_6M]
        p12 = close.iloc[-WEEKS_12M]
        m3  = momentum(current, p3)
        m6  = momentum(current, p6)
        m12 = momentum(current, p12)
        avg = np.mean([m3, m6, m12])

        # Trend filter (10M MA)
        ma10  = close.rolling(MA_10_MONTHS).mean().iloc[-1]
        above = current > ma10

        # 52-week annualised volatility
        vol_52w = annualised_vol(close.iloc[-WEEKS_12M - 1:])

        asset_data[asset_name] = {
            "ticker":   ticker,
            "current":  current,
            "p3": p3, "p6": p6, "p12": p12,
            "m3": m3, "m6": m6, "m12": m12,
            "avg_mom":  avg,
            "ma10":     ma10,
            "above_ma": above,
            "vol_52w":  vol_52w,
        }
        logging.info(f"{asset_name}: OK  avg_mom={avg:.4f}  above_ma={above}  vol={vol_52w:.3f}")

    except Exception as e:
        logging.error(f"{asset_name}: ERROR -> {e}")

if not asset_data:
    print("❌ Yahoo returned no usable data. Check run.log")
    sys.exit(1)

# =========================================================
# COMPOSITE SCORE & ALLOCATION
# =========================================================

# Step 1: vol z-score across the universe
vols     = np.array([d["vol_52w"] for d in asset_data.values()])
vol_mean = vols.mean()
vol_std  = vols.std() if vols.std() > 0 else 1.0
for name, d in asset_data.items():
    d["vol_z"] = (d["vol_52w"] - vol_mean) / vol_std

# Step 2: composite score = momentum × vol_adj × trend_adj
for name, d in asset_data.items():
    vol_adj   = max(1.0 - VOL_PENALTY * d["vol_z"], VOL_ADJ_FLOOR)
    trend_adj = 1.0 if d["above_ma"] else TREND_PENALTY
    d["score"] = d["avg_mom"] * vol_adj * trend_adj
    logging.info(
        f"{name}: avg_mom={d['avg_mom']:+.4f}  vol_z={d['vol_z']:+.2f}  "
        f"vol_adj={vol_adj:.3f}  trend_adj={trend_adj}  score={d['score']:+.4f}"
    )

# Step 3: shift scores so all are positive, normalise to weights
scores = np.array([d["score"] for d in asset_data.values()])
shift  = abs(scores.min()) + 0.01 if scores.min() <= 0 else 0.0
total  = sum(d["score"] + shift for d in asset_data.values())
alloc  = {name: (d["score"] + shift) / total for name, d in asset_data.items()}

# Sanity check
total_alloc = sum(alloc.values())
assert abs(total_alloc - 1.0) < 1e-6, f"Weights sum to {total_alloc:.6f}, not 1.0"
logging.info(f"Final allocation: { {n: f'{w:.2%}' for n, w in alloc.items()} }")

# =========================================================
# BUILD OUTPUT ROWS
# =========================================================

rows = []

for name in sorted(asset_data, key=lambda n: -alloc[n]):
    d  = asset_data[name]
    w  = alloc[name]
    rk = sorted(alloc, key=lambda n: -alloc[n]).index(name) + 1

    rows.append([
        name,
        d["ticker"],
        round(d["p3"],      2),
        round(d["p6"],      2),
        round(d["p12"],     2),
        round(d["current"], 2),
        round(d["m3"],      4),
        round(d["m6"],      4),
        round(d["m12"],     4),
        round(d["avg_mom"], 4),
        "Y" if d["above_ma"] else "N",
        round(d["vol_52w"],      4),
        round(d["vol_z"],        2),
        round(d["score"],        4),
        rk,
        round(w, 4),
    ])

columns = [
    "Asset Name", "Ticker",
    "Price 3M Ago", "Price 6M Ago", "Price 12M Ago", "Current Price",
    "3M Momentum", "6M Momentum", "12M Momentum", "Avg Momentum",
    "Above 10M MA?", "52W Ann. Vol", "Vol Z-Score", "Composite Score",
    "Rank", "Allocation Weight",
]

df_out = pd.DataFrame(rows, columns=columns)

# =========================================================
# EXPORT TO EXCEL WITH FORMATTING
# =========================================================

df_out.to_excel(OUTPUT_FILE, index=False, sheet_name="Momentum_Signal")

wb = load_workbook(OUTPUT_FILE)
ws = wb["Momentum_Signal"]

# ── Styles ────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="2C2C2A")
HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
ALLOC_FILL   = PatternFill("solid", fgColor="EAF3DE")   # light green
NOALLOC_FILL = PatternFill("solid", fgColor="F9F9F9")
YES_FONT  = Font(name="Arial", size=10, color="3B6D11", bold=True)
NO_FONT   = Font(name="Arial", size=10, color="993C1D")
POS_FONT  = Font(name="Arial", size=10, color="0F6E56")
NEG_FONT  = Font(name="Arial", size=10, color="993C1D")
BORDER    = Border(
    bottom=Side(style="thin", color="E0E0E0"),
    right=Side(style="thin",  color="E0E0E0"),
)

# Column widths
col_widths = [22, 14, 13, 13, 13, 13, 11, 11, 11, 13, 12, 11, 11, 14, 6, 16]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Header row
for cell in ws[1]:
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 32

# Data rows
num_cols  = {3, 4, 5, 6}    # price cols
mom_cols  = {7, 8, 9, 10}   # momentum cols
bool_cols = {11}             # Above 10M MA
vol_col   = 12               # 52W Ann. Vol
volz_col  = 13               # Vol Z-Score
score_col = 14               # Composite Score
rank_col  = 15
alloc_col = 16

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    has_alloc = row[alloc_col - 1].value and row[alloc_col - 1].value > 0
    row_fill  = ALLOC_FILL if has_alloc else NOALLOC_FILL

    for cell in row:
        cell.font      = BODY_FONT
        cell.fill      = row_fill
        cell.border    = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        c = cell.column

        if c == 1:  # asset name — left align
            cell.alignment = Alignment(horizontal="left", vertical="center")

        if c in num_cols:
            cell.number_format = "#,##0.00"

        if c in mom_cols and isinstance(cell.value, float):
            cell.number_format = "0.00%"
            cell.font = POS_FONT if cell.value >= 0 else NEG_FONT

        if c in bool_cols:
            if cell.value == "Y":
                cell.font = YES_FONT
            elif cell.value == "N":
                cell.font = NO_FONT

        if c == vol_col:
            cell.number_format = "0.00%"

        if c == volz_col and isinstance(cell.value, float):
            cell.number_format = "0.00"
            cell.font = NEG_FONT if cell.value > 0.5 else BODY_FONT

        if c == score_col and isinstance(cell.value, float):
            cell.number_format = "0.0000"
            cell.font = POS_FONT if cell.value >= 0 else NEG_FONT

        if c == alloc_col and isinstance(cell.value, float):
            cell.number_format = "0.00%"
            cell.font = Font(name="Arial", size=10, bold=True, color="0F6E56")

# Freeze header
ws.freeze_panes = "A2"

# ── Add a summary block to the right ──────────────────────
sum_col = len(columns) + 2   # leave a gap column
ws.cell(1, sum_col).value = "ALLOCATION SUMMARY"
ws.cell(1, sum_col).font  = HDR_FONT
ws.cell(1, sum_col).fill  = HDR_FILL
ws.cell(1, sum_col).alignment = Alignment(horizontal="center")
ws.column_dimensions[get_column_letter(sum_col)].width = 22
ws.column_dimensions[get_column_letter(sum_col + 1)].width = 12

ws.cell(2, sum_col).value = "Asset"
ws.cell(2, sum_col + 1).value = "Weight"
for c in [sum_col, sum_col + 1]:
    ws.cell(2, c).font = Font(name="Arial", bold=True, size=10)
    ws.cell(2, c).fill = PatternFill("solid", fgColor="EEEDFE")

for i, (name, w) in enumerate(sorted(alloc.items(), key=lambda x: -x[1]), start=3):
    ws.cell(i, sum_col).value     = name
    ws.cell(i, sum_col).font      = BODY_FONT
    ws.cell(i, sum_col + 1).value = w
    ws.cell(i, sum_col + 1).number_format = "0.00%"
    ws.cell(i, sum_col + 1).font  = Font(name="Arial", size=10, bold=True, color="0F6E56")

total_row = 3 + len(alloc)
ws.cell(total_row, sum_col).value = "TOTAL"
ws.cell(total_row, sum_col).font  = Font(name="Arial", bold=True, size=10)
ws.cell(total_row, sum_col + 1).value = sum(alloc.values())
ws.cell(total_row, sum_col + 1).number_format = "0.00%"
ws.cell(total_row, sum_col + 1).font  = Font(name="Arial", bold=True, size=10)

wb.save(OUTPUT_FILE)

# =========================================================
# CONSOLE SUMMARY
# =========================================================

print("\n✅  Run complete")
print(f"    Assets fetched : {len(asset_data)}")
print(f"\n{'Asset':<25} {'Alloc':>7}  {'Avg Mom':>9}  {'Score':>8}  {'Vol Z':>6}  {'Above MA':>9}")
print("─" * 76)
for name in sorted(alloc, key=lambda n: -alloc[n]):
    d  = asset_data[name]
    ma = "Yes" if d["above_ma"] else " No"
    print(f"  {name:<23} {alloc[name]:>7.2%}  {d['avg_mom']:>+9.2%}  "
          f"{d['score']:>+8.4f}  {d['vol_z']:>+6.2f}  {ma:>9}")

print(f"\n    Output: {OUTPUT_FILE}")
print(f"    Log   : {LOG_FILE}")

logging.info(f"Allocation: { {n: f'{w:.2%}' for n, w in alloc.items()} }")
logging.info("===== RUN SUCCESS =====")